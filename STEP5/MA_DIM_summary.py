import openpyxl
import openpyxl.styles as styles
import os

# Path to the directory containing the .txt files
txt_files_dir = '/home/guest/Traineeship/Data/MA_DIM_genes/Results/'

# Load the MA_DIM_genes.xlsx file
genes_workbook = openpyxl.load_workbook('/home/guest/Traineeship/Scripts/STEP4/MA_DIM_genes.xlsx')
genes_sheet = genes_workbook.active

# Create a new workbook for MA_DIM_Results.xlsx
results_workbook = openpyxl.Workbook()
results_sheet = results_workbook.active

# Set the header values in the results sheet
results_sheet['A1'] = 'Locus tag'
results_sheet['B1'] = 'Species'

# Get the Rv numbers from MA_DIM_genes.xlsx
rv_numbers = [str(cell.value) for cell in genes_sheet['B'] if cell.value is not None][1:]

# Write the Rv numbers in column A of the results sheet
for i, rv_number in enumerate(rv_numbers):
    results_sheet.cell(row=i + 3, column=1, value=rv_number)

# Get the species names from the .txt files
species_names = [filename[:-11] for filename in os.listdir(txt_files_dir) if filename.endswith('_Result.txt')]

# Write the species names in row 2, starting from column B
for i, species_name in enumerate(species_names):
    results_sheet.cell(row=2, column=i + 2, value=species_name)

# Creating cell style --> no value in cell, means light red background color
red_fill = styles.PatternFill(fill_type="solid", fgColor="F7D1D5")
no_style = styles.NamedStyle(name="no_value")
no_style.fill = red_fill

# Iterate over the .txt files and update the corresponding cells in the results sheet
for i, rv_number in enumerate(rv_numbers):
    for j, species_name in enumerate(species_names):
        txt_filename = os.path.join(txt_files_dir, f"{species_name}_Result.txt")
        if os.path.isfile(txt_filename):
            with open(txt_filename, 'r') as txt_file:
                if rv_number in txt_file.read():
                    results_sheet.cell(row=i + 3, column=j + 2, value='yes')
                else:
                    results_sheet.cell(row=i + 3, column=j + 2, value='no')
                    results_sheet.cell(row=i + 3, column=j + 2).style = no_style

# Save the results workbook as MA_DIM_Results.xlsx
results_workbook.save('/home/guest/Traineeship/Scripts/STEP5/MA_DIM_Results.xlsx')
